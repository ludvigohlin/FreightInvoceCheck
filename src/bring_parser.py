"""Bring-specific parsers for PDF invoice headers and Excel specification lines."""

from __future__ import annotations

import re
import shutil
import tempfile
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from src import config
from src.processing_logger import ProcessingLogger
from src.utils import parse_swedish_number, parse_date, safe_float, infer_country_from_postal_code

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ── Bring description parsing ─────────────────────────────────────────────────
# Format: "4027 Business Parcel Bulk, Fuel fee (332)"
#     or: "9999 Business Parcel Bulk (332)"
#     or: "332 Business Parcel Bulk, Fuel fee"   (no trailing route code)
#     or: "332 Business Parcel Bulk"
_DESC_PATTERN = re.compile(
    r"^(\d+)\s+([^,(]+?)(?:,\s*([^(]+?))?\s*(?:\((\d+)\))?\s*$"
)

# Bring invoice number prefix (typically starts with 40)
_BRING_INV_PATTERN = re.compile(r"faktura(?:nr\.?|nummer)?\s+(\d{8,12})", re.IGNORECASE)
_BRING_DATE_PATTERN = re.compile(r"(\d{4}-\d{2}-\d{2})")


@dataclass
class BringInvoiceHeader:
    run_id: str
    processed_timestamp: str
    carrier: str = "Bring"
    invoice_number: str = ""
    invoice_date: str = ""
    due_date: str = ""
    customer_number: str = ""
    customer_reference: str = ""
    period_from: str = ""
    period_to: str = ""
    currency: str = ""
    total_ex_vat: Optional[float] = None
    vat_amount: Optional[float] = None
    total_inc_vat: Optional[float] = None
    source_file: str = ""
    document_type: str = "Invoice"
    reconciliation_status: str = ""
    error_message: str = ""

    def to_dict(self) -> dict:
        return {
            "run_id": self.run_id,
            "processed_timestamp": self.processed_timestamp,
            "carrier": self.carrier,
            "invoice_number": self.invoice_number,
            "invoice_date": self.invoice_date,
            "due_date": self.due_date,
            "customer_number": self.customer_number,
            "customer_reference": self.customer_reference,
            "period_from": self.period_from,
            "period_to": self.period_to,
            "currency": self.currency,
            "total_ex_vat": self.total_ex_vat,
            "vat_amount": self.vat_amount,
            "total_inc_vat": self.total_inc_vat,
            "source_file": self.source_file,
            "document_type": self.document_type,
            "reconciliation_status": self.reconciliation_status,
            "error_message": self.error_message,
        }


@dataclass
class BringInvoiceLine:
    run_id: str
    processed_timestamp: str
    invoice_number: str
    carrier: str = "Bring"
    source_file: str = ""
    line_no: int = 0
    # shipment identifiers
    shipment_number: str = ""
    package_number: str = ""
    from_postal: str = ""
    to_postal: str = ""
    from_country: str = "SE"
    to_country: str = "NO"
    # article / service
    article_number: str = ""
    service_code: str = ""
    service_name_raw: str = ""
    # parsed from description
    base_service_name: str = ""
    surcharge_name_raw: str = ""
    route_code: str = ""
    # classification
    service_category: str = "Unknown"
    surcharge_category: str = ""
    line_type: str = "Unknown"
    classified_by: str = "Rules"
    classification_confidence: float = 1.0
    manual_review_required: bool = False
    # quantities / pricing
    quantity: float = 1.0
    unit: str = "St"
    gross_price: Optional[float] = None
    unit_price: Optional[float] = None   # = avtalspris (contract price per unit)
    amount: Optional[float] = None       # = avtalspris total (this is what gets invoiced)
    discount_percent: Optional[float] = None
    weight_kg: Optional[float] = None
    vat_type: str = ""
    received_datetime: str = ""

    def to_dict(self) -> dict:
        return {
            "run_id": self.run_id,
            "processed_timestamp": self.processed_timestamp,
            "carrier": self.carrier,
            "invoice_number": self.invoice_number,
            "source_file": self.source_file,
            "line_no": self.line_no,
            "article_number": self.article_number,
            "service_code": self.service_code,
            "service_name_raw": self.service_name_raw,
            "service_category": self.service_category,
            "from_country": self.from_country,
            "to_country": self.to_country,
            "quantity": self.quantity,
            "unit": self.unit,
            "unit_price": self.unit_price,
            "discount_percent": self.discount_percent,
            "vat_type": self.vat_type,
            "amount": self.amount,
            "line_type": self.line_type,
            "classified_by": self.classified_by,
            "classification_confidence": self.classification_confidence,
            "manual_review_required": self.manual_review_required,
        }

    def to_surcharge_dict(self) -> dict:
        return {
            "run_id": self.run_id,
            "processed_timestamp": self.processed_timestamp,
            "carrier": self.carrier,
            "invoice_number": self.invoice_number,
            "source_file": self.source_file,
            "line_no": self.line_no,
            "surcharge_raw": self.surcharge_name_raw,
            "surcharge_category": self.surcharge_category,
            "service_name_raw": self.service_name_raw,
            "quantity": self.quantity,
            "unit_price": self.unit_price,
            "amount": self.amount,
            "related_service_category": self.service_category,
            "classified_by": self.classified_by,
            "classification_confidence": self.classification_confidence,
            "manual_review_required": self.manual_review_required,
        }


# ── PDF header extraction ─────────────────────────────────────────────────────

def parse_bring_pdf_header(
    file_path: Path,
    run_id: str,
    logger: ProcessingLogger,
) -> Optional[BringInvoiceHeader]:
    """
    Extract invoice header fields from a Bring PDF invoice.
    Returns BringInvoiceHeader or None on failure.
    """
    from src.pdf_utils import extract_text_from_pdf

    try:
        text = extract_text_from_pdf(file_path)
    except Exception as e:
        logger.error("BringParser", f"Cannot read PDF: {e}", file_name=file_path.name, error=e)
        return None

    ts = datetime.now().isoformat(timespec="seconds")
    header = BringInvoiceHeader(
        run_id=run_id,
        processed_timestamp=ts,
        source_file=file_path.name,
    )

    # Invoice number
    m = re.search(r"faktura(?:nr\.?|nummer)?\s+(\d{8,12})", text, re.IGNORECASE)
    if m:
        header.invoice_number = m.group(1)

    # Invoice date
    m = re.search(r"fakturadatum\s+(\d{4}-\d{2}-\d{2})", text, re.IGNORECASE)
    if m:
        header.invoice_date = m.group(1)

    # Due date — may show as "Förfallodatum" or garbled: "F.rfallodatum"
    m = re.search(r"f.rfallodatum\s+(\d{4}-\d{2}-\d{2})", text, re.IGNORECASE)
    if m:
        header.due_date = m.group(1)

    # Customer number
    m = re.search(r"kundnummer\s+(\d+)", text, re.IGNORECASE)
    if m:
        header.customer_number = m.group(1)

    # Customer reference (Kundreferens) — may be empty; exclude IBAN/BIC-style values
    m = re.search(r"kundreferens\s+([^\n]{1,40})", text, re.IGNORECASE)
    if m:
        val = m.group(1).strip()
        # Exclude IBAN (SE...), BIC (ESSE...), and pure numeric bank references
        upper_val = val.upper()
        if (val
                and not upper_val.startswith("IBAN")
                and not upper_val.startswith("SE29")  # IBAN starting with country code
                and not re.match(r"^[A-Z]{4}[A-Z]{2}", val)  # BIC/SWIFT
                and not val.isdigit()):
            header.customer_reference = val

    # Order period
    m = re.search(
        r"orderperiod\s+(\d{4}-\d{2}-\d{2})\s*[-–]\s*(\d{4}-\d{2}-\d{2})",
        text, re.IGNORECASE,
    )
    if m:
        header.period_from = m.group(1)
        header.period_to = m.group(2)

    # Currency
    m = re.search(r"valuta\s+([A-Z]{3})", text, re.IGNORECASE)
    if m:
        header.currency = m.group(1).upper()

    # Total ex VAT: "Summa exkl. moms  15 556,25"
    m = re.search(r"summa\s+exkl\.?\s*moms\s+([\d\s\xa0]+,\d{2})", text, re.IGNORECASE)
    if m:
        header.total_ex_vat = parse_swedish_number(m.group(1))

    # Total incl VAT: "Summa inkl. moms  15 556,25"
    m = re.search(r"summa\s+inkl\.?\s*moms\s+([\d\s\xa0]+,\d{2})", text, re.IGNORECASE)
    if m:
        header.total_inc_vat = parse_swedish_number(m.group(1))

    # Fallback: "Att betala  15 556,25"
    if header.total_inc_vat is None:
        m = re.search(r"att\s+betala\s+([\d\s\xa0]+,\d{2})", text, re.IGNORECASE)
        if m:
            header.total_inc_vat = parse_swedish_number(m.group(1))

    # Derive VAT = total_inc_vat - total_ex_vat (avoids greedy regex issues in the VAT table)
    if header.total_ex_vat is not None and header.total_inc_vat is not None:
        header.vat_amount = round(header.total_inc_vat - header.total_ex_vat, 2)

    logger.info(
        "BringParser",
        f"PDF header extracted: invoice={header.invoice_number}, "
        f"total_ex_vat={header.total_ex_vat}, date={header.invoice_date}",
        file_name=file_path.name,
    )
    return header


# ── Excel specification parsing ───────────────────────────────────────────────

# Expected column names in the Bring specification Excel (Swedish originals)
_BRING_EXCEL_EXPECTED_HEADERS = [
    "skickat", "postnummer", "försändelse",  # försändelse
    "artikelnummer", "beskrivning", "produkt", "avtalspris",
]

# Normalized column name aliases → standard field name
_COLUMN_ALIASES = {
    "skickat_fr_n_postnummer": "from_postal",
    "skickat_fran_postnummer": "from_postal",
    "postnummer_leveransadress": "to_postal",
    "f_rs_ndelsenummer": "shipment_number",
    "fors_ndelsenummer": "shipment_number",
    "kollinummer": "package_number",
    "artikelnummer": "article_number",
    "beskrivning": "description",
    "produkt": "product",
    "fraktber_knad_vikt_kg": "weight_kg",
    "fraktber_knad_vikt__kg": "weight_kg",
    "antal_kolli": "quantity",
    "mottaget_av_bring_datum_och_tid": "received_datetime",
    "moms": "vat_pct",
    "moms___": "vat_pct",
    "bruttopris": "gross_price",
    "avtalspris": "amount",
    "fakturanummer": "invoice_number_col",
    "fakturadatum": "invoice_date_col",
    "drivmedelstill_gg": "fuel_surcharge_col",
    "svaveltill_gg": "sulphur_surcharge_col",
}


def _normalize_col(raw: str) -> str:
    """Normalize a column name for alias lookup."""
    import re as _re
    s = str(raw or "").strip().lower()
    s = _re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def _map_columns(headers: list[str]) -> dict[str, int]:
    """
    Map standard field names to column indices based on normalized header names.
    Returns dict: field_name → column_index
    """
    mapping = {}
    for i, h in enumerate(headers):
        norm = _normalize_col(h)
        standard = _COLUMN_ALIASES.get(norm)
        if standard:
            mapping[standard] = i
        else:
            # Try partial match for robustness
            for alias, std in _COLUMN_ALIASES.items():
                if alias in norm or norm in alias:
                    if std not in mapping:
                        mapping[std] = i
    return mapping


def _parse_description(desc: str) -> dict:
    """
    Parse a Bring description like "4027 Business Parcel Bulk, Fuel fee (332)".
    Returns dict with: service_code, base_service, surcharge_raw, route_code.
    """
    if not desc:
        return {}
    m = _DESC_PATTERN.match(str(desc).strip())
    if not m:
        return {"service_name_raw": str(desc).strip()}
    service_code = m.group(1)
    base_service = m.group(2).strip()
    surcharge_raw = m.group(3).strip() if m.group(3) else ""
    route_code = m.group(4)
    return {
        "service_code": service_code,
        "base_service_name": base_service,
        "surcharge_name_raw": surcharge_raw,
        "route_code": route_code,
        "service_name_raw": str(desc).strip(),
    }


def _classify_line(
    service_code: str,
    base_service: str,
    surcharge_raw: str,
    service_mapping: dict,
    surcharge_mapping: dict,
) -> dict:
    """
    Deterministically classify a line into service_category, surcharge_category, line_type.
    Returns classification dict.
    """
    service_cats = service_mapping.get("service_categories", {})
    line_type_by_code = service_mapping.get("line_type_by_article_prefix", {})
    surcharge_cats = surcharge_mapping.get("surcharge_categories", {})

    base_lower = base_service.lower()
    surcharge_lower = surcharge_raw.lower()

    # Determine line_type from article code first
    line_type = line_type_by_code.get(service_code, "Unknown")

    # If no article code match, use presence of surcharge name
    if line_type == "Unknown":
        line_type = "Surcharge" if surcharge_raw else "BaseFreight"

    # Determine service_category from base service name
    service_category = "Unknown"
    for cat, keywords in service_cats.items():
        if any(kw.lower() in base_lower for kw in keywords):
            service_category = cat
            break

    # Determine surcharge_category from surcharge name
    surcharge_category = ""
    if surcharge_raw:
        for cat, keywords in surcharge_cats.items():
            if any(kw.lower() in surcharge_lower for kw in keywords):
                surcharge_category = cat
                break
        if not surcharge_category:
            surcharge_category = "Unknown"

    manual_review = service_category == "Unknown" or (surcharge_raw and surcharge_category == "Unknown")

    return {
        "service_category": service_category,
        "surcharge_category": surcharge_category,
        "line_type": line_type,
        "classified_by": "Rules",
        "classification_confidence": 0.5 if manual_review else 1.0,
        "manual_review_required": manual_review,
    }


def parse_bring_excel_specification(
    file_path: Path,
    run_id: str,
    logger: ProcessingLogger,
    service_mapping: dict = None,
    surcharge_mapping: dict = None,
) -> tuple[Optional[BringInvoiceHeader], List[BringInvoiceLine]]:
    """
    Parse a Bring Excel specification file.
    Returns (header_from_excel, list_of_lines).
    The header from Excel has summary totals which are used for reconciliation.
    """
    import openpyxl

    if service_mapping is None:
        service_mapping = config.load_service_mapping()
    if surcharge_mapping is None:
        surcharge_mapping = config.load_surcharge_mapping()

    ts = datetime.now().isoformat(timespec="seconds")

    # Copy to a temp file first — avoids OneDrive/Windows file-lock errors
    try:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copy2(str(file_path), tmp.name)
        wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
    except Exception as e:
        logger.error("BringParser", f"Cannot open Excel: {e}", file_name=file_path.name, error=e)
        return None, []
    finally:
        try:
            Path(tmp.name).unlink(missing_ok=True)
        except Exception:
            pass

    ws = wb.active
    spec_header = BringInvoiceHeader(
        run_id=run_id,
        processed_timestamp=ts,
        source_file=file_path.name,
        document_type="Specification",
    )

    # Read all rows into memory once — ws.cell() on a read_only workbook is O(n²),
    # iter_rows(values_only=True) streams the XML in one pass.
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    # ── Pass 1: Extract summary header fields from top rows ───────────────────
    header_row_idx = None
    for row_idx, row_vals_all in enumerate(all_rows[:30], start=1):
        row_str = " ".join(str(v) for v in row_vals_all if v is not None).lower()
        # Use first 2 cols for summary value extraction
        col1 = row_vals_all[0] if row_vals_all else None
        col2 = row_vals_all[1] if len(row_vals_all) > 1 else None

        if "fakturanummer" in row_str or "f\xf6r fakturanummer" in row_str or "for fakturanummer" in row_str:
            # Extract invoice number from any cell in this row
            for v in row_vals_all:
                if v:
                    m = re.search(r"(\d{8,12})", str(v))
                    if m:
                        spec_header.invoice_number = m.group(1)
                        break

        if "isicom" in row_str:
            # Customer reference row: "IsiCom AB Isicom Norge (20023849191)"
            for v in row_vals_all:
                if v:
                    m = re.search(r"\((\d+)\)", str(v))
                    if m:
                        spec_header.customer_number = m.group(1)
                        break

        if "total utan moms" in row_str:
            if col2 is not None:
                spec_header.total_ex_vat = parse_swedish_number(str(col2))

        if "total moms" in row_str and "total utan moms" not in row_str:
            if col2 is not None:
                spec_header.vat_amount = parse_swedish_number(str(col2))

        if "faktura totalsumma" in row_str or ("totalsumma" in row_str and "fakturanummer" not in row_str):
            if col2 is not None:
                spec_header.total_inc_vat = parse_swedish_number(str(col2))

        # Detect header row: contains both "Artikelnummer"/"Beskrivning" AND "Avtalspris"
        has_service_col = "artikelnummer" in row_str or "beskrivning" in row_str
        has_price_col = "avtalspris" in row_str or "bruttopris" in row_str
        if has_service_col and has_price_col:
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        logger.warning(
            "BringParser",
            "Could not find data header row in Excel specification",
            file_name=file_path.name,
        )
        return spec_header, []

    # ── Read column headers ───────────────────────────────────────────────────
    raw_headers = all_rows[header_row_idx - 1]  # all_rows is 0-indexed
    col_map = _map_columns([str(h) if h else "" for h in raw_headers])

    logger.info(
        "BringParser",
        f"Excel header found at row {header_row_idx}; columns mapped: {list(col_map.keys())}",
        file_name=file_path.name,
    )

    # ── Pass 2: Parse data rows ───────────────────────────────────────────────
    lines: List[BringInvoiceLine] = []
    line_no = 0

    for row_vals in all_rows[header_row_idx:]:  # all_rows is 0-indexed; header_row_idx+1 row = index header_row_idx
        # Skip empty rows
        if all(v is None for v in row_vals):
            continue

        def get(field: str):
            idx = col_map.get(field)
            return row_vals[idx] if idx is not None and idx < len(row_vals) else None

        description = str(get("description") or "").strip()
        article_num = str(get("article_number") or "").strip()

        if not description and not article_num:
            continue  # skip rows without meaningful content

        line_no += 1
        desc_parsed = _parse_description(description)

        # Determine from/to country from postal codes
        from_postal = str(get("from_postal") or "").strip()
        to_postal = str(get("to_postal") or "").strip()
        from_country = infer_country_from_postal_code(from_postal) if from_postal else "SE"
        to_country = infer_country_from_postal_code(to_postal) if to_postal else "NO"

        # Amount = Avtalspris (contracted/agreed price — what is actually charged)
        amount_raw = get("amount")
        amount = safe_float(amount_raw) if amount_raw is not None else None

        # Gross price
        gross_raw = get("gross_price")
        gross_price = safe_float(gross_raw) if gross_raw is not None else None

        # Discount percent — derive from gross vs contract price
        discount_pct = None
        if gross_price and amount is not None and gross_price > 0 and amount != gross_price:
            discount_pct = round((1 - amount / gross_price) * 100, 2)

        # Quantity and weight
        qty = safe_float(get("quantity")) or 1.0
        weight = safe_float(get("weight_kg"))

        # VAT type — Bring export invoices have 0% VAT; represent as "Export"
        vat_pct_raw = get("vat_pct")
        vat_type = ""
        if vat_pct_raw is not None:
            vat_float = safe_float(vat_pct_raw)
            if vat_float == 0.0:
                vat_type = "Export"
            elif vat_float is not None:
                vat_type = f"{int(vat_float)}%"

        # Received datetime
        recv_raw = get("received_datetime")
        received_dt = ""
        if recv_raw:
            if isinstance(recv_raw, datetime):
                received_dt = recv_raw.isoformat(timespec="seconds")
            else:
                received_dt = str(recv_raw)

        # Invoice number from data row (should match header)
        inv_from_row = str(get("invoice_number_col") or "").strip()
        invoice_number = inv_from_row or spec_header.invoice_number

        # Classification
        classification = _classify_line(
            service_code=desc_parsed.get("service_code", ""),
            base_service=desc_parsed.get("base_service_name", description),
            surcharge_raw=desc_parsed.get("surcharge_name_raw", ""),
            service_mapping=service_mapping,
            surcharge_mapping=surcharge_mapping,
        )

        # Unit price = amount / qty if qty > 1, else = amount
        unit_price = round(amount / qty, 4) if (amount is not None and qty and qty > 0) else amount

        line = BringInvoiceLine(
            run_id=run_id,
            processed_timestamp=ts,
            invoice_number=invoice_number,
            source_file=file_path.name,
            line_no=line_no,
            shipment_number=str(get("shipment_number") or "").strip(),
            package_number=str(get("package_number") or "").strip(),
            from_postal=from_postal,
            to_postal=to_postal,
            from_country=from_country,
            to_country=to_country,
            article_number=article_num,
            service_code=desc_parsed.get("service_code", ""),
            service_name_raw=description,
            base_service_name=desc_parsed.get("base_service_name", ""),
            surcharge_name_raw=desc_parsed.get("surcharge_name_raw", ""),
            route_code=desc_parsed.get("route_code", ""),
            service_category=classification["service_category"],
            surcharge_category=classification["surcharge_category"],
            line_type=classification["line_type"],
            classified_by=classification["classified_by"],
            classification_confidence=classification["classification_confidence"],
            manual_review_required=classification["manual_review_required"],
            quantity=qty,
            unit="St",
            gross_price=gross_price,
            unit_price=unit_price,
            amount=amount,
            discount_percent=discount_pct,
            weight_kg=weight,
            vat_type=vat_type,
            received_datetime=received_dt,
        )
        lines.append(line)

    logger.info(
        "BringParser",
        f"Excel parsed: {len(lines)} lines, spec_total_ex_vat={spec_header.total_ex_vat}",
        file_name=file_path.name,
    )
    return spec_header, lines
