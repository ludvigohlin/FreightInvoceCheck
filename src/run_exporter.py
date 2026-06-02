"""Generate per-run Excel invoice approval report and HTML archive."""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src import config
from src.processing_logger import ProcessingLogger


# ── Colour palette ─────────────────────────────────────────────────────────────
_WHITE    = "FFFFFFFF"
_GRAY_ROW = "FFF5F7FB"
_GRAY_BRD = "FFE0E0E0"

_HDR_FILL  = PatternFill("solid", fgColor="FF1565C0")
_HDR_FONT  = Font(bold=True, color=_WHITE, size=10)
_BOLD      = Font(bold=True)
_ALT_FILL  = PatternFill("solid", fgColor=_GRAY_ROW)
_WFILL     = PatternFill("solid", fgColor=_WHITE)
_SEC_FILL  = PatternFill("solid", fgColor="FFE3F0FF")
_SEC_FONT  = Font(bold=True, size=9, color="FF1565C0")

_STATUS_FILL = {
    "OK":      PatternFill("solid", fgColor="FFD4EDDA"),
    "Warning": PatternFill("solid", fgColor="FFFFF3CD"),
    "Error":   PatternFill("solid", fgColor="FFF8D7DA"),
}
_STATUS_FONT = {
    "OK":      Font(bold=True, color="FF155724"),
    "Warning": Font(bold=True, color="FF856404"),
    "Error":   Font(bold=True, color="FF842029"),
}
_SEV_FILL = {
    "Error":   PatternFill("solid", fgColor="FFF8D7DA"),
    "Warning": PatternFill("solid", fgColor="FFFFFDE7"),
    "Info":    PatternFill("solid", fgColor="FFE3F2FD"),
}
_SEV_FONT = {
    "Error":   Font(bold=True, color="FF842029"),
    "Warning": Font(bold=True, color="FF856404"),
    "Info":    Font(bold=True, color="FF1565C0"),
}

_BSIDE  = Side(style="thin", color=_GRAY_BRD)
_BORDER = Border(left=_BSIDE, right=_BSIDE, top=_BSIDE, bottom=_BSIDE)
_STATUS_ICON = {"OK": "✓", "Warning": "⚠", "Error": "✗"}


def _fmt(val: Any) -> str:
    if val is None: return ""
    if isinstance(val, float): return f"{val:,.2f}"
    return str(val)


def _auto_width(ws, max_width: int = 60, padding: int = 3) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_len + padding, max_width
        )


def _write_header_row(ws, row: int, values: list, num_cols: int | None = None) -> None:
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _HDR_FONT; c.fill = _HDR_FILL; c.border = _BORDER
    if num_cols:
        for col in range(len(values) + 1, num_cols + 1):
            ws.cell(row=row, column=col).fill = _HDR_FILL
            ws.cell(row=row, column=col).border = _BORDER


def _section_header(ws, row: int, title: str, num_cols: int) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(num_cols)}{row}")
    c = ws.cell(row=row, column=1, value=title)
    c.font = _SEC_FONT; c.fill = _SEC_FILL; c.alignment = Alignment(indent=1)
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).border = _BORDER
        ws.cell(row=row, column=col).fill = _SEC_FILL


def _kv(ws, row: int, key: str, val: Any, num_cols: int = 5,
        fmt: str | None = None, status: str | None = None) -> None:
    k = ws.cell(row=row, column=1, value=key)
    k.font = _BOLD; k.fill = _WFILL; k.border = _BORDER
    k.alignment = Alignment(indent=1)
    v = ws.cell(row=row, column=2, value=val)
    v.fill = _WFILL; v.border = _BORDER
    if fmt:
        v.number_format = fmt
        v.alignment = Alignment(horizontal="right")
    if status:
        v.fill = _STATUS_FILL.get(status, _WFILL)
        v.font = _STATUS_FONT.get(status, _BOLD)
    for col in range(3, num_cols + 1):
        ws.cell(row=row, column=col).fill = _WFILL
        ws.cell(row=row, column=col).border = _BORDER


def _total_row(ws, row: int, values: list) -> None:
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color=_WHITE)
        c.fill = _HDR_FILL; c.border = _BORDER
        if isinstance(val, float):
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
        elif isinstance(val, int):
            c.alignment = Alignment(horizontal="right")


def _style_rows(ws, first: int, last: int, ncols: int,
                status_col: int | None = None) -> None:
    for i, r in enumerate(range(first, last + 1)):
        base = _ALT_FILL if i % 2 else _WFILL
        skey = None
        if status_col:
            raw = str(ws.cell(row=r, column=status_col).value or "")
            for k in ("Error", "Warning", "OK"):
                if k in raw:
                    skey = k; break
        rfill = _STATUS_FILL.get(skey, base) if skey else base
        for col in range(1, ncols + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = rfill; cell.border = _BORDER
            if status_col and col == status_col and skey:
                cell.fill = _STATUS_FILL[skey]
                cell.font = _STATUS_FONT[skey]


# ── Service/surcharge aggregation ─────────────────────────────────────────────

def _agg_service(lines) -> list[dict]:
    """Aggregate base freight lines by service category."""
    data: dict = defaultdict(lambda: {"shipments": set(), "lines": 0, "total": 0.0})
    for ln in lines:
        if getattr(ln, "line_type", "") != "BaseFreight":
            continue
        cat = getattr(ln, "service_category", None) or "Unknown"
        ship = getattr(ln, "shipment_number", None)
        if ship:
            data[cat]["shipments"].add(ship)
        data[cat]["lines"] += 1
        data[cat]["total"] += getattr(ln, "amount", 0.0) or 0.0

    result = []
    for cat, d in sorted(data.items(), key=lambda x: -x[1]["total"]):
        total = round(d["total"], 2)
        n_ship = len(d["shipments"]) or d["lines"]
        result.append({
            "category": cat,
            "shipments": n_ship,
            "total": total,
            "avg": round(total / n_ship, 2) if n_ship else 0.0,
        })
    return result


def _agg_surcharge(lines) -> list[dict]:
    """Aggregate surcharge lines by surcharge category."""
    data: dict = defaultdict(float)
    for ln in lines:
        if getattr(ln, "line_type", "") != "Surcharge":
            continue
        cat = getattr(ln, "surcharge_category", None) or "Unknown"
        data[cat] += getattr(ln, "amount", 0.0) or 0.0
    return [
        {"category": cat, "total": round(amt, 2)}
        for cat, amt in sorted(data.items(), key=lambda x: -x[1])
    ]


# ── Sheet 1: Approval Overview ────────────────────────────────────────────────

def _sheet_overview(ws, run_id: str, headers, checks, anomalies, payload: dict,
                    missing_bring: list) -> None:
    NCOLS = 7
    ws.sheet_properties.tabColor = "1565C0"
    ws.freeze_panes = "A2"

    # Title banner
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    t = ws["A1"]
    t.value = f"Freight Invoice Report  —  {datetime.now().strftime('%Y-%m-%d')}"
    t.font = Font(bold=True, size=14, color=_WHITE)
    t.fill = _HDR_FILL
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t.border = _BORDER
    ws.row_dimensions[1].height = 28
    for col in range(2, NCOLS + 1):
        ws.cell(row=1, column=col).fill = _HDR_FILL
        ws.cell(row=1, column=col).border = _BORDER

    row = 3
    _section_header(ws, row, "Run Information", NCOLS); row += 1
    _kv(ws, row, "Run ID", run_id, NCOLS); row += 1
    _kv(ws, row, "Generated", datetime.now().strftime("%Y-%m-%d %H:%M"), NCOLS); row += 1
    _kv(ws, row, "Files Scanned", payload.get("total_files_scanned", 0), NCOLS); row += 1

    c = payload.get("check_counts", {})
    ok_n, warn_n, err_n = c.get("OK", 0), c.get("Warning", 0), c.get("Error", 0)
    overall = "Error" if err_n else ("Warning" if warn_n else "OK")
    row += 1

    # Invoice status table
    _section_header(ws, row, "Invoice Status", NCOLS); row += 1
    cols = ["Carrier", "Invoice #", "Date", "Amount ex VAT (SEK)",
            "Reconciliation", "Issues", "Status"]
    _write_header_row(ws, row, cols); row += 1

    checks_by_inv: dict = defaultdict(list)
    for c in checks:
        checks_by_inv[c.invoice_number].append(c)
    anomalies_by_inv: dict = defaultdict(list)
    for a in (anomalies or []):
        anomalies_by_inv[a.invoice_number].append(a)

    grand_total = 0.0
    first_data_row = row
    for h in headers:
        inv_checks = checks_by_inv.get(h.invoice_number, [])
        n_err = sum(1 for c in inv_checks if c.severity == "Error")
        n_warn = sum(1 for c in inv_checks if c.severity == "Warning")
        n_anom = len(anomalies_by_inv.get(h.invoice_number, []))
        issue_count = n_err + n_warn + n_anom

        recon_chk = next(
            (c for c in inv_checks if c.check_name in
             ("PDFTotalVsExcelSummary", "LineSumVsHeaderTotal")), None
        )
        recon_txt = ""
        if recon_chk:
            icon = _STATUS_ICON.get(recon_chk.status, "?")
            recon_txt = f"{icon} {recon_chk.status}"

        inv_status = "Error" if n_err else ("Warning" if (n_warn or n_anom) else "OK")
        status_txt = f"{_STATUS_ICON[inv_status]} {'OK' if inv_status == 'OK' else inv_status}"
        amt = h.total_ex_vat or 0.0
        grand_total += amt

        vals = [h.carrier, h.invoice_number, str(h.invoice_date or ""),
                amt, recon_txt, issue_count if issue_count else "—", status_txt]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(indent=1)
            if col == 4:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col == 6:
                cell.alignment = Alignment(horizontal="center")
        # colour status cell
        sc = ws.cell(row=row, column=7)
        sc.fill = _STATUS_FILL.get(inv_status, _WFILL)
        sc.font = _STATUS_FONT.get(inv_status, _BOLD)
        row += 1

    _style_rows(ws, first_data_row, row - 1, NCOLS, status_col=7)
    # Total row
    _total_row(ws, row, ["", "TOTAL", "", round(grand_total, 2), "", "", ""]); row += 1
    row += 1

    # Carrier totals
    _section_header(ws, row, "Totals by Carrier", NCOLS); row += 1
    carrier_totals = payload.get("carrier_totals", {})
    for carrier, info in sorted(carrier_totals.items()):
        _kv(ws, row, f"{carrier} ({info['invoices']} invoice(s))",
            round(info["total_ex_vat"], 2), NCOLS, fmt="#,##0.00"); row += 1
    _kv(ws, row, "GRAND TOTAL", round(grand_total, 2), NCOLS, fmt="#,##0.00"); row += 1
    row += 1

    # Action items
    issue_checks = [c for c in checks if c.severity in ("Error", "Warning")]
    warn_anomalies = [a for a in (anomalies or []) if a.severity in ("Error", "Warning")]
    if issue_checks or warn_anomalies or missing_bring:
        _section_header(ws, row, "⚠ Action Items", NCOLS); row += 1
        for c in sorted(issue_checks, key=lambda x: 0 if x.severity == "Error" else 1):
            icon = "✗" if c.severity == "Error" else "⚠"
            txt = f"{icon}  {c.carrier} {c.invoice_number} — {c.check_name}: {c.message}"
            if c.claude_explanation:
                txt += f"  |  AI: {c.claude_explanation}"
            cell = ws.cell(row=row, column=1, value=txt)
            cell.fill = _STATUS_FILL.get(c.severity, _WFILL)
            cell.font = _STATUS_FONT.get(c.severity, Font())
            cell.border = _BORDER
            cell.alignment = Alignment(indent=1, wrap_text=True)
            ws.merge_cells(f"A{row}:{get_column_letter(NCOLS)}{row}")
            for col in range(2, NCOLS + 1):
                ws.cell(row=row, column=col).border = _BORDER
                ws.cell(row=row, column=col).fill = _STATUS_FILL.get(c.severity, _WFILL)
            ws.row_dimensions[row].height = 30
            row += 1
        for a in warn_anomalies:
            icon = "✗" if a.severity == "Error" else "⚠"
            txt = f"{icon}  {a.carrier} {a.invoice_number} — {a.anomaly_type}: {a.description}"
            if a.claude_explanation:
                txt += f"  |  AI: {a.claude_explanation}"
            cell = ws.cell(row=row, column=1, value=txt)
            cell.fill = _SEV_FILL.get(a.severity, _WFILL)
            cell.font = _SEV_FONT.get(a.severity, Font())
            cell.border = _BORDER
            cell.alignment = Alignment(indent=1, wrap_text=True)
            ws.merge_cells(f"A{row}:{get_column_letter(NCOLS)}{row}")
            for col in range(2, NCOLS + 1):
                ws.cell(row=row, column=col).border = _BORDER
                ws.cell(row=row, column=col).fill = _SEV_FILL.get(a.severity, _WFILL)
            ws.row_dimensions[row].height = 30
            row += 1
        for m in missing_bring:
            txt = f"⚠  Bring {m.get('invoice_number', '')} — {m.get('message', 'Incomplete document set')}"
            cell = ws.cell(row=row, column=1, value=txt)
            cell.fill = _STATUS_FILL["Warning"]; cell.font = _STATUS_FONT["Warning"]
            cell.border = _BORDER; cell.alignment = Alignment(indent=1)
            ws.merge_cells(f"A{row}:{get_column_letter(NCOLS)}{row}")
            for col in range(2, NCOLS + 1):
                ws.cell(row=row, column=col).border = _BORDER
                ws.cell(row=row, column=col).fill = _STATUS_FILL["Warning"]
            row += 1
        row += 1

    # Approval sign-off
    _section_header(ws, row, "Approval Sign-off", NCOLS); row += 1
    for label in ["Reviewed by", "Date", "Comments"]:
        k = ws.cell(row=row, column=1, value=label)
        k.font = _BOLD; k.fill = _WFILL; k.border = _BORDER; k.alignment = Alignment(indent=1)
        ws.merge_cells(f"B{row}:{get_column_letter(NCOLS)}{row}")
        for col in range(2, NCOLS + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFAFAFA")
            ws.cell(row=row, column=col).border = _BORDER
        ws.row_dimensions[row].height = 22
        row += 1

    # Column widths
    widths = [14, 16, 12, 20, 18, 9, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Per-invoice sheet ─────────────────────────────────────────────────────────

def _sheet_invoice_detail(ws, header, lines, checks, anomalies) -> None:
    carrier = header.carrier
    inv_num = header.invoice_number
    NCOLS = 6

    recon_check = next(
        (c for c in checks if c.check_name in
         ("PDFTotalVsExcelSummary", "LineSumVsHeaderTotal")), None
    )
    inv_status = "OK"
    if any(c.severity == "Error" for c in checks):
        inv_status = "Error"
    elif any(c.severity == "Warning" for c in checks) or any(
        a.severity in ("Error", "Warning") for a in anomalies
    ):
        inv_status = "Warning"

    status_colors = {
        "OK": ("FF1565C0", "FFE3F0FF"),
        "Warning": ("FFE65100", "FFFFF3E0"),
        "Error": ("FFC62828", "FFFFEBEE"),
    }
    fg, bg = status_colors.get(inv_status, ("FF1565C0", "FFE3F0FF"))

    tab_colors = {"OK": "1565C0", "Warning": "E65100", "Error": "C62828"}
    ws.sheet_properties.tabColor = tab_colors.get(inv_status, "1565C0")

    # ── Invoice header banner ─────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    t = ws["A1"]
    t.value = f"{carrier}  —  Invoice {inv_num}"
    t.font = Font(bold=True, size=13, color=_WHITE)
    t.fill = PatternFill("solid", fgColor=fg)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t.border = _BORDER
    ws.row_dimensions[1].height = 26
    for col in range(2, NCOLS + 1):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=fg)
        ws.cell(row=1, column=col).border = _BORDER

    row = 3
    info_rows = [
        ("Invoice Date", str(header.invoice_date or "—")),
        ("Total ex VAT", header.total_ex_vat),
        ("Currency", header.currency or "SEK"),
        ("Status", f"{_STATUS_ICON.get(inv_status, '?')} {inv_status}"),
    ]
    if recon_check:
        info_rows.append(("Reconciliation", f"{_STATUS_ICON.get(recon_check.status, '?')} {recon_check.message}"))

    for key, val in info_rows:
        k = ws.cell(row=row, column=1, value=key)
        k.font = _BOLD; k.fill = _WFILL; k.border = _BORDER; k.alignment = Alignment(indent=1)
        v = ws.cell(row=row, column=2, value=val)
        v.fill = _WFILL; v.border = _BORDER
        if key == "Total ex VAT" and isinstance(val, float):
            v.number_format = "#,##0.00"
            v.alignment = Alignment(horizontal="right")
        if key == "Status":
            v.fill = _STATUS_FILL.get(inv_status, _WFILL)
            v.font = _STATUS_FONT.get(inv_status, _BOLD)
        for col in range(3, NCOLS + 1):
            ws.cell(row=row, column=col).fill = _WFILL
            ws.cell(row=row, column=col).border = _BORDER
        row += 1
    row += 1

    # ── Service Breakdown ─────────────────────────────────────────────────────
    svc_breakdown = _agg_service(lines)
    base_total = sum(s["total"] for s in svc_breakdown)

    _section_header(ws, row, "Freight Cost by Service", NCOLS); row += 1
    _write_header_row(ws, row,
                      ["Service Type", "Shipments", "Total (SEK)",
                       "Avg / Shipment (SEK)", "% of Invoice", ""]); row += 1
    first = row
    for s in svc_breakdown:
        pct = s["total"] / base_total * 100 if base_total else 0
        vals = [s["category"], s["shipments"], s["total"],
                s["avg"], pct / 100, ""]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = _BORDER; c.alignment = Alignment(indent=1)
            if col in (3, 4):
                c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
            elif col == 5:
                c.number_format = "0.0%"; c.alignment = Alignment(horizontal="right")
            elif col == 2:
                c.alignment = Alignment(horizontal="right")
        row += 1
    _style_rows(ws, first, row - 1, NCOLS)
    grand_ship = sum(s["shipments"] for s in svc_breakdown)
    _total_row(ws, row, ["TOTAL", grand_ship, round(base_total, 2), "", "", ""]); row += 1
    row += 1

    # ── Surcharge Breakdown ───────────────────────────────────────────────────
    surcharge_breakdown = _agg_surcharge(lines)
    surcharge_total = sum(s["total"] for s in surcharge_breakdown)
    invoice_total = (header.total_ex_vat or 0.0)

    _section_header(ws, row, "Surcharges", NCOLS); row += 1
    _write_header_row(ws, row,
                      ["Surcharge Type", "Amount (SEK)", "% of Surcharges",
                       "% of Invoice", "", ""]); row += 1
    first = row
    for s in surcharge_breakdown:
        pct_sc = s["total"] / surcharge_total * 100 if surcharge_total else 0
        pct_inv = s["total"] / invoice_total * 100 if invoice_total else 0
        vals = [s["category"], s["total"], pct_sc / 100, pct_inv / 100, "", ""]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = _BORDER; c.alignment = Alignment(indent=1)
            if col == 2:
                c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
            elif col in (3, 4):
                c.number_format = "0.0%"; c.alignment = Alignment(horizontal="right")
        row += 1
    if surcharge_breakdown:
        _style_rows(ws, first, row - 1, NCOLS)
        sc_pct_inv = surcharge_total / invoice_total * 100 if invoice_total else 0
        _total_row(ws, row, ["TOTAL SURCHARGES", round(surcharge_total, 2),
                              "", sc_pct_inv / 100, "", ""])
        ws.cell(row=row, column=4).number_format = "0.0%"
        row += 1
    row += 1

    # ── Validation ────────────────────────────────────────────────────────────
    key_checks = [c for c in checks if c.severity in ("Error", "Warning")]
    # Always include the main reconciliation check
    if recon_check and recon_check not in key_checks:
        key_checks.insert(0, recon_check)

    if key_checks:
        _section_header(ws, row, "Validation Issues", NCOLS); row += 1
        _write_header_row(ws, row, ["Check", "Status", "Message", "AI Explanation", "", ""])
        row += 1
        first = row
        for c in key_checks:
            icon = _STATUS_ICON.get(c.status, "?")
            vals = [c.check_name, f"{icon} {c.status}", c.message,
                    c.claude_explanation or "", "", ""]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = _BORDER
                cell.alignment = Alignment(indent=1, wrap_text=(col in (3, 4)))
            ws.cell(row=row, column=2).fill = _STATUS_FILL.get(c.status, _WFILL)
            ws.cell(row=row, column=2).font = _STATUS_FONT.get(c.status, Font())
            ws.row_dimensions[row].height = 32
            row += 1
        _style_rows(ws, first, row - 1, NCOLS, status_col=2)
        row += 1

    # ── Anomalies ─────────────────────────────────────────────────────────────
    if anomalies:
        _section_header(ws, row, "Anomalies", NCOLS); row += 1
        _write_header_row(ws, row, ["Severity", "Type", "Description",
                                     "AI Explanation", "Suggested Action", ""])
        row += 1
        first = row
        for a in sorted(anomalies, key=lambda x: {"Error": 0, "Warning": 1, "Info": 2}.get(x.severity, 9)):
            vals = [f"{_STATUS_ICON.get(a.severity, '?')} {a.severity}",
                    a.anomaly_type, a.description,
                    a.claude_explanation or "", a.suggested_action or "", ""]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = _BORDER
                cell.alignment = Alignment(indent=1, wrap_text=(col in (3, 4, 5)))
            ws.cell(row=row, column=1).fill = _SEV_FILL.get(a.severity, _WFILL)
            ws.cell(row=row, column=1).font = _SEV_FONT.get(a.severity, Font())
            ws.row_dimensions[row].height = 40
            row += 1
        _style_rows(ws, first, row - 1, NCOLS)
        row += 1

    # ── Sign-off ──────────────────────────────────────────────────────────────
    _section_header(ws, row, "Approval", NCOLS); row += 1
    for label in ["Approved by", "Date", "Comments"]:
        k = ws.cell(row=row, column=1, value=label)
        k.font = _BOLD; k.fill = _WFILL; k.border = _BORDER; k.alignment = Alignment(indent=1)
        ws.merge_cells(f"B{row}:{get_column_letter(NCOLS)}{row}")
        for col in range(2, NCOLS + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFAFAFA")
            ws.cell(row=row, column=col).border = _BORDER
        ws.row_dimensions[row].height = 22
        row += 1

    # Column widths
    widths = [22, 14, 22, 40, 32, 4]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Pending files sheet ───────────────────────────────────────────────────────

def _sheet_pending(ws, missing: list) -> None:
    ws.sheet_properties.tabColor = "F9A825"
    _AMBER = PatternFill("solid", fgColor="FFF9A825")

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "⚠ Pending — Incomplete Invoice Sets"
    t.font = Font(bold=True, size=11, color=_WHITE)
    t.fill = _AMBER
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t.border = _BORDER
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = _AMBER
        ws.cell(row=1, column=col).border = _BORDER
    ws.row_dimensions[1].height = 22

    _write_header_row(ws, 2, ["Invoice #", "Carrier", "Status", "Note"])
    ws.freeze_panes = "A3"

    for i, m in enumerate(missing, start=3):
        found = m.get("found_file", "")
        missing_f = m.get("missing_file", "")
        note = f"{found} received — {missing_f} not yet found"
        vals = [m.get("invoice_number", ""), "Bring", "Incomplete", note]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor="FFFFFDE7")
            c.border = _BORDER; c.alignment = Alignment(indent=1)
        ws.cell(row=i, column=3).font = Font(bold=True, color="FF856404")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 65


# ── Main Excel writer ─────────────────────────────────────────────────────────

def _write_excel(
    path: Path,
    run_id: str,
    payload: dict,
    headers,
    lines,
    checks,
    anomalies=None,
    missing_bring=None,
    all_lines_dict: dict | None = None,
) -> None:
    wb = openpyxl.Workbook()
    anomalies = anomalies or []
    missing_bring = missing_bring or []

    # Group checks and anomalies by invoice number for fast lookup
    checks_by_inv: dict = defaultdict(list)
    for c in checks:
        checks_by_inv[c.invoice_number].append(c)
    anom_by_inv: dict = defaultdict(list)
    for a in anomalies:
        anom_by_inv[a.invoice_number].append(a)

    # Sheet 1: Overview
    ws1 = wb.active
    ws1.title = "Overview"
    _sheet_overview(ws1, run_id, headers, checks, anomalies, payload, missing_bring)

    # One sheet per invoice
    for h in headers:
        inv_lines = []
        if all_lines_dict:
            inv_lines = all_lines_dict.get((h.carrier, h.invoice_number), [])
        else:
            inv_lines = [
                ln for ln in lines
                if getattr(ln, "invoice_number", None) == h.invoice_number
                or getattr(ln, "invoice_number_col", None) == h.invoice_number
            ]

        inv_checks = checks_by_inv.get(h.invoice_number, [])
        inv_anom = anom_by_inv.get(h.invoice_number, [])

        # Sheet name: max 31 chars, must be unique
        sheet_name = f"{h.carrier[:4]} {h.invoice_number}"[-31:]
        ws = wb.create_sheet(title=sheet_name)
        _sheet_invoice_detail(ws, h, inv_lines, inv_checks, inv_anom)

    if missing_bring:
        ws_p = wb.create_sheet("Pending")
        _sheet_pending(ws_p, missing_bring)

    wb.save(path)


# ── HTML archive (lightweight, not emailed) ───────────────────────────────────

def _write_html_archive(path: Path, run_id: str, payload: dict) -> None:
    """Write a minimal HTML archive record for this run."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    c = payload.get("check_counts", {})
    ok_n, warn_n, err_n = c.get("OK", 0), c.get("Warning", 0), c.get("Error", 0)
    overall = "Error" if err_n else ("Warning" if warn_n else "OK")
    total = sum(v.get("total_ex_vat", 0) for v in payload.get("carrier_totals", {}).values())

    content = (
        f"<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<title>Freight Invoice Control {run_id}</title></head><body>"
        f"<h2>Freight Invoice Control</h2>"
        f"<p>Run: {run_id} | Generated: {now} | Status: {overall}</p>"
        f"<p>Invoices: {len(payload.get('invoices', []))} | "
        f"Total: {total:,.2f} SEK | Checks: {ok_n} OK / {warn_n} Warning / {err_n} Error</p>"
        f"<p>See attached Excel report for full details.</p>"
        f"</body></html>"
    )
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)


# ── Public entry point ────────────────────────────────────────────────────────

def write_run_export(
    run_id: str,
    payload: dict,
    all_invoice_headers,
    all_invoice_lines,
    all_checks,
    logger: ProcessingLogger,
    ai_summary: str | None = None,
    anomalies=None,
    missing_bring: list | None = None,
    all_lines_dict: dict | None = None,
) -> None:
    """Write XLSX approval report into For_Email/ and HTML archive into Summaries/."""
    config.FOR_EMAIL_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = config.FOR_EMAIL_DIR / f"summary_{run_id}.xlsx"
    html_path = config.SUMMARIES_DIR / f"summary_{run_id}.html"

    _write_excel(
        xlsx_path, run_id, payload,
        all_invoice_headers, all_invoice_lines, all_checks,
        anomalies=anomalies or [],
        missing_bring=missing_bring or [],
        all_lines_dict=all_lines_dict,
    )
    _write_html_archive(html_path, run_id, payload)

    logger.info("RunExporter",
                f"Power Automate file: {xlsx_path.name} | HTML archive: {html_path.name}")


def write_missing_file_alert(
    run_id: str,
    missing: list[dict],
    logger: ProcessingLogger,
) -> None:
    """Write a For_Email XLSX alert when a Bring invoice is missing its PDF or Excel pair."""
    if not missing:
        return

    config.FOR_EMAIL_DIR.mkdir(parents=True, exist_ok=True)
    path = config.FOR_EMAIL_DIR / f"alert_missing_files_{run_id}.xlsx"
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    n = len(missing)

    _RED = PatternFill("solid", fgColor="FFC62828")
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Alert"
    ws1.sheet_properties.tabColor = "C62828"

    ws1.merge_cells("A1:C1")
    t = ws1["A1"]
    t.value = "⚠ ACTION REQUIRED — Missing Invoice File(s)"
    t.font = Font(bold=True, size=13, color=_WHITE)
    t.fill = _RED
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t.border = _BORDER
    for col in (2, 3):
        ws1.cell(row=1, column=col).fill = _RED
        ws1.cell(row=1, column=col).border = _BORDER
    ws1.row_dimensions[1].height = 26

    for row, (key, val) in enumerate([
        ("Run ID", run_id), ("Generated", now_str), ("Incomplete invoices", n),
        ("Action", "Retrieve missing file, place in 00_Inbox, re-run."),
    ], start=3):
        k = ws1.cell(row=row, column=1, value=key)
        k.font = _BOLD; k.fill = _WFILL; k.border = _BORDER; k.alignment = Alignment(indent=1)
        v = ws1.cell(row=row, column=2, value=val)
        v.fill = _WFILL; v.border = _BORDER
        ws1.cell(row=row, column=3).border = _BORDER
        ws1.cell(row=row, column=3).fill = _WFILL

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 65

    ws2 = wb.create_sheet("Incomplete Invoices")
    ws2.sheet_properties.tabColor = "C62828"
    _write_header_row(ws2, 1, ["Invoice #", "Missing File", "Received File", "Details"])
    ws2.freeze_panes = "A2"
    for i, m in enumerate(missing, start=2):
        ws2.cell(row=i, column=1, value=m.get("invoice_number", ""))
        ws2.cell(row=i, column=2, value=m.get("missing_file", "")).font = Font(color="FFC62828")
        ws2.cell(row=i, column=3, value=m.get("found_file", "")).font = Font(color="FF2E7D32")
        ws2.cell(row=i, column=4, value=m.get("message", ""))
    _auto_width(ws2)

    wb.save(path)
    logger.warning("RunExporter",
                   f"Missing-file alert written: {path.name} ({n} incomplete invoice(s))")
