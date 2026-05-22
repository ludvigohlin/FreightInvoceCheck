"""Excel utility functions — find header rows, normalize column names, load data."""

from __future__ import annotations

from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl

from src.utils import normalize_column_name


def find_header_row(
    ws,
    expected_keywords: List[str],
    max_scan_rows: int = 30,
) -> Optional[int]:
    """
    Scan the first max_scan_rows rows to find a row that looks like a header.
    Returns the 1-based row index, or None if not found.

    A row is considered a header if it contains at least 3 of the expected keywords
    (case-insensitive, partial match allowed).
    """
    keywords_lower = [k.lower() for k in expected_keywords]
    for row_idx in range(1, max_scan_rows + 1):
        row_values = [
            str(ws.cell(row=row_idx, column=c).value or "").lower()
            for c in range(1, ws.max_column + 1)
        ]
        matches = sum(
            1 for kw in keywords_lower
            if any(kw in cell for cell in row_values)
        )
        if matches >= 3:
            return row_idx
    return None


def load_sheet_as_dicts(
    file_path: Path,
    sheet_name: Optional[str] = None,
    header_row: Optional[int] = None,
    expected_keywords: Optional[List[str]] = None,
    max_scan_rows: int = 30,
) -> Tuple[List[dict], List[str]]:
    """
    Load an Excel sheet into a list of dicts.
    Auto-detects header row if header_row is None and expected_keywords are given.

    Returns (rows, column_names).
    """
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    if header_row is None and expected_keywords:
        header_row = find_header_row(ws, expected_keywords, max_scan_rows)

    if header_row is None:
        wb.close()
        return [], []

    # Read header names
    raw_headers = [
        ws.cell(row=header_row, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    normalized_headers = [normalize_column_name(str(h)) if h is not None else f"col_{i}"
                          for i, h in enumerate(raw_headers)]

    # Read data rows
    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = [
            ws.cell(row=row_idx, column=c).value
            for c in range(1, ws.max_column + 1)
        ]
        # Skip completely empty rows
        if all(v is None for v in row_values):
            continue
        row_dict = dict(zip(normalized_headers, row_values))
        rows.append(row_dict)

    wb.close()
    return rows, normalized_headers


def read_cell_range(ws, start_row: int, end_row: int) -> List[Tuple]:
    """Read all rows between start_row and end_row (inclusive), return list of tuples."""
    result = []
    for row_idx in range(start_row, end_row + 1):
        row_vals = tuple(
            ws.cell(row=row_idx, column=c).value
            for c in range(1, ws.max_column + 1)
        )
        result.append(row_vals)
    return result


def find_cell_with_text(ws, keyword: str, max_rows: int = 20) -> Optional[Tuple[int, int]]:
    """
    Find the first cell containing keyword (case-insensitive).
    Returns (row, col) 1-based, or None.
    """
    kw_lower = keyword.lower()
    for row_idx in range(1, max_rows + 1):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and kw_lower in str(val).lower():
                return (row_idx, col_idx)
    return None
