"""
freight_summary.py
──────────────────
Builds the 4-tab freight invoice attestation Excel from structured invoice data.

USAGE
─────
    from freight_summary import build_summary, SummaryInput, Invoice, Service, Surcharge, Anomaly

    data = SummaryInput(
        run_id="20260602_191024_dc327c83",
        generated="2026-06-02 19:11",
        files_scanned=8,
        invoices=[...],
        services=[...],
        unallocated=[...],
        surcharges=[...],
        anomalies=[...],
    )
    build_summary(data, "output/summary_20260602.xlsx")

DATA CONTRACT
─────────────
Your invoice parser must populate the dataclasses below.
All amounts are ex. VAT, in SEK (or whatever currency — just be consistent).

TABS PRODUCED
─────────────
  1. Attest          — verdict banner, AI-review block, carrier rollup, invoice table, sign-off
  2. Kostnad per tjänst — base freight by service per carrier, with tillägg bridge row
  3. Tillägg         — surcharges per carrier; unknown charges flagged red
  4. Avvikelser      — anomalies sorted by severity; Regel vs AI source column

DEPENDENCIES
────────────
    pip install openpyxl
"""

from __future__ import annotations
from dataclasses import dataclass, field
from typing import Literal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src import config


# ══════════════════════════════════════════════════════════════════════════════
# DATA CONTRACTS  (populate these from your invoice parser)
# ══════════════════════════════════════════════════════════════════════════════

Status   = Literal["OK", "Warning", "Error"]
Severity = Literal["Error", "Warning", "Info"]
Source   = Literal["Regel", "AI"]


@dataclass
class Invoice:
    carrier: str                    # e.g. "Bring", "PostNord"
    number: str                     # invoice number as string
    date: str                       # "YYYY-MM-DD"
    total_ex_vat: float             # header total from the invoice
    recon_status: Status            # result of line-sum vs header check
    recon_message: str              # human-readable one-liner, shown in "Att kolla"
    n_anomalies: int                # count of anomalies tied to this invoice
    status: Status                  # overall invoice status (worst of recon + anomalies)


@dataclass
class Service:
    """Base freight costs, excluding surcharges."""
    carrier: str
    service_name: str               # e.g. "Kolli (parcel)", "Pall (pallet)"
    shipments: int                  # number of distinct sändningar (deliveries)
    total_ex_vat: float
    packages: int = 0               # total kolli/lines (for Bring: > shipments when multi-kolli orders)


@dataclass
class Unallocated:
    """
    Amount from an invoice that cannot be broken into service lines
    (e.g. unparsed supplement invoice, reconciliation gap).
    """
    carrier: str
    label: str                      # shown as service name
    amount: float


@dataclass
class Surcharge:
    carrier: str
    name: str                       # surcharge description
    amount: float
    is_fuel: bool                   # True  → categorised as fuel surcharge
    is_unknown: bool                # True  → AI/rules could not classify → flagged red


@dataclass
class Anomaly:
    severity: Severity
    carrier: str
    invoice_number: str
    anomaly_type: str               # short label, e.g. "Hög radkostnad"
    description: str                # what was found
    suggested_action: str           # what the reviewer should do
    ai_explanation: str             # AI reasoning (empty string if source == "Regel")
    source: Source                  # "Regel" = caught by rule engine, "AI" = AI escalation
    country: str = ""               # destination country code, set for NonNordicDestination
    amount: float = 0.0             # total cost in invoice currency, set for NonNordicDestination
    shipment_count: int = 0         # shipment count, set for NonNordicDestination


@dataclass
class ServiceSurcharge:
    """Surcharge amount tied to a specific service type (for per-service breakdown)."""
    carrier: str
    service_name: str               # matches Service.service_name
    surcharge_name: str             # e.g. "Bränsle (fuel)"
    amount: float
    is_fuel: bool
    is_unknown: bool


@dataclass
class Pending:
    """Bring invoice where only one of the two required files has arrived."""
    invoice_number: str
    missing_file: str
    found_file: str


@dataclass
class SummaryInput:
    run_id: str
    generated: str                  # "YYYY-MM-DD HH:MM"
    files_scanned: int
    invoices: list[Invoice]
    services: list[Service]
    surcharges: list[Surcharge]
    anomalies: list[Anomaly]
    unallocated: list[Unallocated] = field(default_factory=list)
    service_surcharges: list[ServiceSurcharge] = field(default_factory=list)
    carrier_currency: dict[str, str] = field(default_factory=dict)  # carrier -> "SEK"/"NOK"/etc.
    pending: list[Pending] = field(default_factory=list)


# ══════════════════════════════════════════════════════════════════════════════
# STYLE CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

_NAVY   = "1F3A5F"; _GREY   = "5B6B7B"; _LINE   = "D9DEE5"
_PEND_F = "E8F4FC"; _PEND_T = "1A5276"
_OK_F   = "E3F1E4"; _OK_T   = "1E6B33"
_WARN_F = "FCEFD6"; _WARN_T = "8A5A00"
_ERR_F  = "FBE0E0"; _ERR_T  = "9B1C1C"
_HDR_F  = "1F3A5F"; _HDR_T  = "FFFFFF"
_SUB_F  = "EEF1F5"
_AI_F   = "F0ECF8"; _AI_T   = "4A2580"
_FONT   = "Arial"

_thin = Side(style="thin", color=_LINE)
_box  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_bot  = Border(bottom=_thin)

_SEK  = '#,##0.00;(#,##0.00);"–"'
_PCT  = '0.0%'
_INT  = '#,##0;(#,##0);"–"'

# Carrier colour map — extend as needed
_CARRIER_FILL = {
    "Bring":    "FFF4E6",
    "PostNord": "E8F0FB",
}
_DEFAULT_CARRIER_FILL = "F5F5F5"


def _carrier_fill(carrier: str) -> str:
    return _CARRIER_FILL.get(carrier, _DEFAULT_CARRIER_FILL)


def _status_style(s: str) -> tuple[str, str]:
    return {
        "OK":      (_OK_F,   _OK_T),
        "Warning": (_WARN_F, _WARN_T),
        "Error":   (_ERR_F,  _ERR_T),
    }.get(s, (_SUB_F, _GREY))


# ══════════════════════════════════════════════════════════════════════════════
# LOW-LEVEL CELL HELPER
# ══════════════════════════════════════════════════════════════════════════════

def _c(ws, r: int, c: int, v=None, *, bold=False, size=10, color="1A1A1A",
       fill=None, align="left", fmt=None, border=None, wrap=False, italic=False):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name=_FONT, size=size, bold=bold, italic=italic, color=color)
    x.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:   x.fill = PatternFill("solid", fgColor=fill)
    if fmt:    x.number_format = fmt
    if border: x.border = border
    return x


def _hdr(ws, r: int, cols: list[str], widths: list[float] | None = None):
    for i, name in enumerate(cols, 1):
        _c(ws, r, i, name, bold=True, color=_HDR_T, fill=_HDR_F,
           align="left" if i == 1 else "center", border=_box)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _scol(n: int) -> str:
    return get_column_letter(n)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def _build_attest(wb: Workbook, d: SummaryInput, carriers: list[str],
                  carrier_totals: dict, carrier_fuel: dict,
                  carrier_other: dict, carrier_frakt: dict) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([20, 16, 13, 18, 14, 12, 12, 48], 1):
        ws.column_dimensions[_scol(i)].width = w

    # Title
    _c(ws, 1, 1, "Fraktfaktura – attestunderlag", bold=True, size=16, color=_NAVY)
    _c(ws, 2, 1,
       f"Körning {d.run_id}  ·  genererad {d.generated}  ·  {d.files_scanned} filer skannade",
       size=9, color=_GREY)

    # Verdict banner
    n_err  = sum(1 for i in d.invoices if i.status == "Error")
    n_warn = sum(1 for i in d.invoices if i.status == "Warning")
    total  = sum(i.total_ex_vat for i in d.invoices)
    if n_err:
        verdict = (f"{len(d.invoices)} fakturor  ·  {total:,.0f} SEK ex moms  ·  "
                   f"{n_err} fel, {n_warn} varningar – kräver granskning före attest")
        vf, vt = _ERR_F, _ERR_T
    elif n_warn:
        verdict = (f"{len(d.invoices)} fakturor  ·  {total:,.0f} SEK ex moms  ·  "
                   f"{n_warn} varningar – kontrollera markerade fakturor")
        vf, vt = _WARN_F, _WARN_T
    elif not d.invoices and d.pending:
        verdict = f"{len(d.pending)} faktura(or) inväntar saknat dokument – se Inväntar dokument nedan"
        vf, vt = _PEND_F, _PEND_T
    else:
        verdict = f"{len(d.invoices)} fakturor  ·  {total:,.0f} SEK ex moms  ·  Inga fel – redo för attest"
        vf, vt = _OK_F, _OK_T
    ws.merge_cells("A4:H4")
    _c(ws, 4, 1, verdict.replace(",", " "), bold=True, size=11, color=vt, fill=vf, border=_box)

    # KPI tiles (4 tiles, 2 cols each)
    r = 6
    tiles = []
    for c in carriers:
        ct  = carrier_totals[c]
        ccy = d.carrier_currency.get(c, "SEK")
        cf  = _carrier_fill(c)
        fuel_amt  = carrier_fuel.get(c, 0.0)
        fuel_pct  = fuel_amt / ct if ct else 0.0
        tiles.append({
            "value": f"{ct:,.0f}",
            "unit":  ccy,
            "label": f"Total {c}",
            "sub":   f"Bränsle {fuel_pct:.0%}",
            "fill":  cf,
        })
    n_anom_issues = sum(1 for a in d.anomalies if a.severity in ("Error", "Warning"))
    n_inv_issues  = n_err + n_warn
    anom_fill = _ERR_F if any(a.severity == "Error" for a in d.anomalies) \
                else _WARN_F if n_anom_issues else _OK_F
    inv_fill  = _ERR_F if n_err else _WARN_F if n_warn else _OK_F
    tiles.append({
        "value": str(n_anom_issues),
        "unit":  "",
        "label": "Avvikelser",
        "sub":   "fel eller varning",
        "fill":  anom_fill,
    })
    tiles.append({
        "value": str(n_inv_issues),
        "unit":  "",
        "label": "Fakturor m. anmärkning",
        "sub":   f"{len(d.invoices)} tot",
        "fill":  inv_fill,
    })

    n_tiles = min(len(tiles), 4)
    for ti, tile in enumerate(tiles[:4]):
        c1 = ti * 2 + 1
        c2 = c1 + 1
        ws.merge_cells(start_row=r,   start_column=c1, end_row=r,   end_column=c2)
        ws.merge_cells(start_row=r+1, start_column=c1, end_row=r+1, end_column=c2)
        ws.merge_cells(start_row=r+2, start_column=c1, end_row=r+2, end_column=c2)
        val_txt = tile["value"] + (f" {tile['unit']}" if tile["unit"] else "")
        _c(ws, r,   c1, val_txt,      bold=True, size=16, color=_NAVY,
           fill=tile["fill"], align="center", border=_box)
        _c(ws, r+1, c1, tile["label"], size=9, color=_NAVY,
           fill=tile["fill"], align="center", border=_box)
        _c(ws, r+2, c1, tile["sub"],   size=8, color=_GREY, italic=True,
           fill=tile["fill"], align="center", border=_box)
    # Fill remaining columns when fewer than 4 tiles
    for fill_col in range(n_tiles * 2 + 1, 9):
        for rr in (r, r + 1, r + 2):
            _c(ws, rr, fill_col, "", border=_box)
    ws.row_dimensions[r].height   = 26
    ws.row_dimensions[r+1].height = 16
    ws.row_dimensions[r+2].height = 14
    r += 4  # 3 tile rows + 1 blank

    # Carrier rollup
    _c(ws, r, 1, "Summa per leverantör", bold=True, size=11, color=_NAVY)
    r += 1
    _hdr(ws, r, ["Leverantör", "Fakturor", "Belopp ex moms",
                 "varav frakt", "varav bränsletillägg", "varav övriga tillägg",
                 "Totalt tillägg", "Andel"])
    r += 1
    crow0 = r
    for c in carriers:
        cf  = _carrier_fill(c)
        ccy = d.carrier_currency.get(c, "SEK")
        _c(ws, r, 1, f"{c} ({ccy})",       bold=True, fill=cf, border=_box)
        _c(ws, r, 2, sum(1 for i in d.invoices if i.carrier == c),
           align="center", fill=cf, border=_box, fmt=_INT)
        _c(ws, r, 3, carrier_totals[c],    align="right", fill=cf, border=_box, fmt=_SEK)
        _c(ws, r, 4, carrier_frakt[c],     align="right", fill=cf, border=_box, fmt=_SEK)
        _c(ws, r, 5, carrier_fuel[c],      align="right", fill=cf, border=_box, fmt=_SEK)
        _c(ws, r, 6, carrier_other[c],     align="right", fill=cf, border=_box, fmt=_SEK)
        _c(ws, r, 7, f"=E{r}+F{r}",       align="right", fill=cf, border=_box, fmt=_SEK)
        _c(ws, r, 8, None,                 align="center", fill=cf, border=_box, fmt=_PCT)
        r += 1
    tot = r
    _c(ws, tot, 1, "TOTAL", bold=True, fill=_SUB_F, border=_box)
    for col, form in [
        (2, f"=SUM(B{crow0}:B{tot-1})"),
        (3, f"=SUM(C{crow0}:C{tot-1})"),
        (4, f"=SUM(D{crow0}:D{tot-1})"),
        (5, f"=SUM(E{crow0}:E{tot-1})"),
        (6, f"=SUM(F{crow0}:F{tot-1})"),
        (7, f"=SUM(G{crow0}:G{tot-1})"),
    ]:
        _c(ws, tot, col, form, bold=True,
           align="right" if col > 2 else "center",
           fill=_SUB_F, border=_box, fmt=_INT if col == 2 else _SEK)
    _c(ws, tot, 8, f"=C{tot}/C{tot}", bold=True, align="center", fill=_SUB_F, border=_box, fmt=_PCT)
    for k, c in enumerate(carriers):
        ws.cell(row=crow0 + k, column=8).value = f"=C{crow0+k}/$C${tot}"
    r = tot + 2

    # ── Non-Nordic destinations ───────────────────────────────────────────────
    # Ordinary shipping countries are SE/NO/DK/FI; anything else is flagged by
    # detect_non_nordic_destinations() and surfaced here so cost is visible at
    # a glance instead of buried as text in the Avvikelser tab.
    non_nordic = [a for a in d.anomalies if a.anomaly_type == "NonNordicDestination"]
    if non_nordic:
        _c(ws, r, 1, "Sändningar utanför Norden (SE/NO/DK/FI)", bold=True, size=11, color=_NAVY)
        r += 1
        _hdr(ws, r, ["Leverantör", "Land", "Sändningar", "Kostnad"])
        r += 1
        agg: dict[tuple[str, str], dict] = {}
        for a in non_nordic:
            key = (a.carrier, a.country or "Unknown")
            e = agg.setdefault(key, {"shipments": 0, "amount": 0.0})
            e["shipments"] += a.shipment_count
            e["amount"] += a.amount
        for (carrier, country), e in sorted(agg.items(), key=lambda kv: -kv[1]["amount"]):
            cf = _carrier_fill(carrier)
            ccy = d.carrier_currency.get(carrier, "SEK")
            _c(ws, r, 1, carrier, fill=cf, border=_box)
            _c(ws, r, 2, country, fill=cf, border=_box, align="center")
            _c(ws, r, 3, e["shipments"], fill=cf, border=_box, align="center", fmt=_INT)
            _c(ws, r, 4, f"{e['amount']:,.2f} {ccy}", fill=cf, border=_box, align="right")
            r += 1
        r += 1

    # Invoice table
    _c(ws, r, 1, "Fakturor – status", bold=True, size=11, color=_NAVY)
    r += 1
    _hdr(ws, r, ["Leverantör", "Faktura #", "Datum", "Belopp ex moms",
                 "Avstämning", "Avvikelser", "Status", "Att kolla"])
    r += 1
    inv_sorted = sorted(d.invoices, key=lambda i: i.date or "")
    for inv in inv_sorted:
        cf = _carrier_fill(inv.carrier)
        sf, st = _status_style(inv.status)
        rf, rt = _status_style(inv.recon_status)
        _c(ws, r, 1, inv.carrier,     fill=cf, border=_box)
        _c(ws, r, 2, inv.number,      border=_box, size=9)
        _c(ws, r, 3, inv.date,        align="center", border=_box)
        _c(ws, r, 4, inv.total_ex_vat, align="right", border=_box, fmt=_SEK)
        _c(ws, r, 5, inv.recon_status, align="center", fill=rf, color=rt, border=_box)
        _c(ws, r, 6, inv.n_anomalies if inv.n_anomalies else None,
           align="center", border=_box, fmt=_INT)
        _c(ws, r, 7, inv.status,      bold=True, align="center", fill=sf, color=st, border=_box)
        note = "" if inv.status == "OK" else inv.recon_message
        _c(ws, r, 8, note,            size=9, color=_GREY, border=_box)
        r += 1
    _c(ws, r, 1, "TOTAL", bold=True, fill=_SUB_F, border=_box)
    _c(ws, r, 2, "", fill=_SUB_F, border=_box)
    _c(ws, r, 3, "", fill=_SUB_F, border=_box)
    _c(ws, r, 4, f"=SUM(D{r-len(d.invoices)}:D{r-1})",
       bold=True, align="right", fill=_SUB_F, border=_box, fmt=_SEK)
    for col in (5, 6, 7, 8):
        _c(ws, r, col, "", fill=_SUB_F, border=_box)
    r += 1
    _tol = config.load_validation_rules().get("reconciliation", {})
    _tol_ok = _tol.get("total_tolerance_ok", 0.01)
    _tol_warn = _tol.get("total_tolerance_warning", 1.00)
    _c(ws, r, 1,
       f"Avstämning \"OK\" tillåter upp till {_tol_ok:.2f} SEK differens; "
       f"upp till {_tol_warn:.2f} SEK visas som \"Warning\"; däröver som \"Error\".",
       size=9, italic=True, color=_GREY)
    r += 2

    # ── Pending (incomplete Bring invoice pairs) ──────────────────────────────
    if d.pending:
        _c(ws, r, 1, "Inväntar dokument", bold=True, size=11, color=_NAVY)
        r += 1
        for col, label in [(1, "Faktura #"), (2, "Mottaget"), (3, "Saknas")]:
            _c(ws, r, col, label, bold=True, color=_HDR_T, fill=_HDR_F,
               align="left" if col == 1 else "center", border=_box)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        _c(ws, r, 4, "Kommentar", bold=True, color=_HDR_T, fill=_HDR_F, border=_box)
        for col in (5, 6, 7, 8):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=_HDR_F)
            ws.cell(row=r, column=col).border = _box
        r += 1
        for p in d.pending:
            _c(ws, r, 1, p.invoice_number, border=_box, fill=_PEND_F)
            _c(ws, r, 2, p.found_file,     border=_box, fill=_PEND_F, size=9)
            _c(ws, r, 3, p.missing_file,   border=_box, fill=_PEND_F, size=9)
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
            _c(ws, r, 4,
               "Lägg saknad fil i 00_Inbox — inkluderas vid nästa körning.",
               border=_box, fill=_PEND_F, size=9, color=_GREY, italic=True)
            for col in (5, 6, 7, 8):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=_PEND_F)
                ws.cell(row=r, column=col).border = _box
            r += 1

    ws.freeze_panes = "A5"


def _build_services(wb: Workbook, d: SummaryInput, carriers: list[str],
                    carrier_totals: dict, carrier_fuel: dict, carrier_other: dict) -> None:
    ws = wb.create_sheet("Kostnad per tjänst")
    ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, "Kostnad per tjänst – per leverantör", bold=True, size=14, color=_NAVY)
    _c(ws, 2, 1, "En rad per tjänstetyp. Tillägg fördelade per tjänst.", size=9, color=_GREY)

    # Columns: Tjänst | Sändningar | Kolli | Basfrakt | Bränsle | Övr. tillägg | Total |
    #          Snitt Basfrakt | Snitt Bränsle | Snitt Övr. tillägg | Snitt/sändning (all "Snitt" per sändning)
    # Currency label is set per carrier section from carrier_currency.
    # Kolli > Sändningar for Bring (multi-kolli shipments); equal for PostNord (1 kolli = 1 sändning).
    WIDTHS = [24, 12, 10, 16, 14, 16, 16, 16, 16, 16, 18]
    NC = 11

    r = 4
    for c in carriers:
        cf  = _carrier_fill(c)
        ct  = carrier_totals[c]
        ccy = d.carrier_currency.get(c, "SEK")
        COLS = ["Tjänst", "Sändningar", "Kolli", f"Basfrakt ({ccy})", f"Bränsle ({ccy})",
                f"Övr. tillägg ({ccy})", f"Total ({ccy})",
                f"Snitt basfrakt ({ccy})", f"Snitt bränsle ({ccy})", f"Snitt övr. tillägg ({ccy})",
                f"Snitt / sändning ({ccy})"]

        # Index service-level surcharges (those that could be attributed to a service type)
        svc_fuel: dict[str, float] = {}
        svc_other: dict[str, float] = {}
        for sc in d.service_surcharges:
            if sc.carrier != c:
                continue
            if sc.is_fuel:
                svc_fuel[sc.service_name]  = svc_fuel.get(sc.service_name, 0.0)  + sc.amount
            else:
                svc_other[sc.service_name] = svc_other.get(sc.service_name, 0.0) + sc.amount

        # Unattributed surcharges (invoice-level, no service linkage — typical for PostNord fuel).
        # Only count a surcharge as "attributed" if it maps to a real service for this carrier;
        # surcharges tied to non-existent service names (e.g. "Oklassificerat") are unattributed.
        existing_svc_names = {s.service_name for s in d.services if s.carrier == c}
        attributed_fuel  = sum(v for k, v in svc_fuel.items()  if k in existing_svc_names)
        attributed_other = sum(v for k, v in svc_other.items() if k in existing_svc_names)
        unattr_fuel  = round(carrier_fuel.get(c, 0.0)  - attributed_fuel,  2)
        unattr_other = round(carrier_other.get(c, 0.0) - attributed_other, 2)

        # Carrier banner
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        _c(ws, r, 1, c, bold=True, size=11, color=_HDR_T, fill=_HDR_F, border=_box)
        for col in range(2, NC + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=_HDR_F)
            ws.cell(row=r, column=col).border = _box
        r += 1

        # Column headers
        _hdr(ws, r, COLS, widths=WIDTHS)
        r += 1

        data_start = r
        for svc in [s for s in d.services if s.carrier == c]:
            fuel   = svc_fuel.get(svc.service_name, 0.0)
            other  = svc_other.get(svc.service_name, 0.0)
            total  = svc.total_ex_vat + fuel + other
            kolli  = svc.packages if svc.packages else svc.shipments
            avg_base  = svc.total_ex_vat / svc.shipments if svc.shipments else 0.0
            avg_fuel  = fuel  / svc.shipments if svc.shipments else 0.0
            avg_other = other / svc.shipments if svc.shipments else 0.0
            avg_s     = total / svc.shipments if svc.shipments else 0.0

            _c(ws, r, 1, svc.service_name, fill=cf, border=_box)
            _c(ws, r, 2, svc.shipments,    align="right", fill=cf, border=_box, fmt=_INT)
            _c(ws, r, 3, kolli if kolli != svc.shipments else None,
               align="right", fill=cf, border=_box, fmt=_INT)
            _c(ws, r, 4, svc.total_ex_vat, align="right", fill=cf, border=_box, fmt=_SEK)
            _c(ws, r, 5, fuel  if fuel  else None, align="right", fill=cf, border=_box, fmt=_SEK)
            _c(ws, r, 6, other if other else None, align="right", fill=cf, border=_box, fmt=_SEK)
            _c(ws, r, 7, total,            align="right", fill=cf, border=_box, fmt=_SEK, bold=True)
            _c(ws, r, 8, avg_base if svc.shipments else None, align="right", fill=cf, border=_box, fmt=_SEK)
            _c(ws, r, 9, avg_fuel if fuel else None, align="right", fill=cf, border=_box, fmt=_SEK)
            _c(ws, r, 10, avg_other if other else None, align="right", fill=cf, border=_box, fmt=_SEK)
            _c(ws, r, 11, avg_s if svc.shipments else None, align="right", fill=cf, border=_box, fmt=_SEK)
            r += 1

        # Invoice-level (unattributed) surcharges — shown as separate row
        if unattr_fuel > 0.01 or unattr_other > 0.01:
            _c(ws, r, 1, "Generella tillägg (fakturanivå)",
               fill=_AI_F, color=_GREY, border=_box, size=9, italic=True)
            for col in (2, 3, 4):
                _c(ws, r, col, None, fill=_AI_F, border=_box)
            _c(ws, r, 5, unattr_fuel  if unattr_fuel  > 0.01 else None,
               align="right", fill=_AI_F, border=_box, fmt=_SEK, italic=True)
            _c(ws, r, 6, unattr_other if unattr_other > 0.01 else None,
               align="right", fill=_AI_F, border=_box, fmt=_SEK, italic=True)
            _c(ws, r, 7, unattr_fuel + unattr_other,
               align="right", fill=_AI_F, border=_box, fmt=_SEK, italic=True)
            for col in (8, 9, 10, 11):
                _c(ws, r, col, None, fill=_AI_F, border=_box)
            r += 1

        # Unallocated rows (e.g. unparsed supplement invoices)
        for u in [u for u in d.unallocated if u.carrier == c]:
            _c(ws, r, 1, u.label, fill=cf, border=_box, size=9, color=_GREY)
            for col in (2, 3, 4, 5, 6):
                _c(ws, r, col, None, border=_box)
            _c(ws, r, 7, u.amount, align="right", fill=cf, border=_box, fmt=_SEK)
            for col in (8, 9, 10, 11):
                _c(ws, r, col, None, border=_box)
            r += 1

        # Totals row
        _c(ws, r, 1, "TOTAL", bold=True, fill=_SUB_F, border=_box)
        _c(ws, r, 2, f"=SUM(B{data_start}:B{r-1})", bold=True, align="right",
           fill=_SUB_F, border=_box, fmt=_INT)
        for col in (3,):
            _c(ws, r, col, None, fill=_SUB_F, border=_box)
        for col, letter in [(4, "D"), (5, "E"), (6, "F"), (7, "G")]:
            _c(ws, r, col, f"=SUM({letter}{data_start}:{letter}{r-1})", bold=True,
               align="right", fill=_SUB_F, border=_box, fmt=_SEK)
        for col, letter in [(8, "D"), (9, "E"), (10, "F"), (11, "G")]:
            _c(ws, r, col, f"=IF(B{r}>0,{letter}{r}/B{r},\"\")", bold=True, align="right",
               fill=_SUB_F, border=_box, fmt=_SEK)
        r += 2

    ws.freeze_panes = "A4"


def _build_surcharges(wb: Workbook, d: SummaryInput, carriers: list[str],
                      carrier_totals: dict) -> None:
    ws = wb.create_sheet("Tillägg")
    ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, "Tillägg (surcharges) – per leverantör", bold=True, size=14, color=_NAVY)
    _c(ws, 2, 1, "Okända/oklassade tillägg markeras i rött – granska innan attest.", size=9, color=_GREY)
    _hdr(ws, 4,
         ["Leverantör", "Tillägg", "Typ", "Belopp ex moms", "Andel av tillägg", "Andel av fakturatotal"],
         widths=[20, 26, 12, 16, 18, 22])
    r = 5
    for c in carriers:
        rows = [s for s in d.surcharges if s.carrier == c]
        start = r
        for s in rows:
            typ = "Bränsle" if s.is_fuel else ("Okänt ⚠" if s.is_unknown else "Övrigt")
            ff  = _ERR_F  if s.is_unknown else (_WARN_F if s.is_fuel else _carrier_fill(c))
            ft  = _ERR_T  if s.is_unknown else (_WARN_T if s.is_fuel else "1A1A1A")
            _c(ws, r, 1, c,      fill=ff, border=_box, color=ft)
            _c(ws, r, 2, s.name, fill=ff, border=_box, color=ft, bold=s.is_unknown)
            _c(ws, r, 3, typ,    fill=ff, border=_box, color=ft, align="center")
            _c(ws, r, 4, s.amount, fill=ff, border=_box, color=ft, align="right", fmt=_SEK)
            _c(ws, r, 5, None,   align="center", border=_box, fmt=_PCT)
            _c(ws, r, 6, f"=D{r}/{carrier_totals[c]}", align="center", border=_box, fmt=_PCT)
            r += 1
        sub = r
        _c(ws, sub, 1, f"{c} – summa tillägg", bold=True, fill=_SUB_F, border=_box)
        _c(ws, sub, 2, "", fill=_SUB_F, border=_box)
        _c(ws, sub, 3, "", fill=_SUB_F, border=_box)
        _c(ws, sub, 4, f"=SUM(D{start}:D{sub-1})", bold=True, align="right", fill=_SUB_F, border=_box, fmt=_SEK)
        _c(ws, sub, 5, f"=D{sub}/D{sub}", bold=True, align="center", fill=_SUB_F, border=_box, fmt=_PCT)
        _c(ws, sub, 6, f"=D{sub}/{carrier_totals[c]}", bold=True, align="center", fill=_SUB_F, border=_box, fmt=_PCT)
        for rr in range(start, sub):
            ws.cell(row=rr, column=5).value = f"=D{rr}/$D${sub}"
        r = sub + 2
    ws.freeze_panes = "A5"


def _build_anomalies(wb: Workbook, d: SummaryInput) -> None:
    ws = wb.create_sheet("Avvikelser")
    ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, "Avvikelser & kontroller att åtgärda", bold=True, size=14, color=_NAVY)
    _c(ws, 2, 1,
       "Källa = Regel: hittad av regelmotorn. Källa = AI: reglerna räckte inte – "
       "AI har bedömt och dessa kräver manuell granskning.",
       size=9, color=_GREY)
    _hdr(ws, 4,
         ["Allvar", "Källa", "Leverantör", "Faktura", "Typ",
          "Beskrivning", "Föreslagen åtgärd", "AI-förklaring"],
         widths=[10, 9, 12, 15, 18, 40, 38, 52])
    sev_order = {"Error": 0, "Warning": 1, "Info": 2}
    sorted_anomalies = sorted(d.anomalies, key=lambda a: sev_order.get(a.severity, 9))
    r = 5
    for a in sorted_anomalies:
        sf, st = _status_style({"Error": "Error", "Warning": "Warning", "Info": "OK"}[a.severity])
        src_f  = _AI_F if a.source == "AI" else _SUB_F
        src_t  = _AI_T if a.source == "AI" else _GREY
        _c(ws, r, 1, a.severity,        bold=True, align="center", fill=sf, color=st, border=_box)
        _c(ws, r, 2, a.source,          bold=(a.source == "AI"), align="center",
           fill=src_f, color=src_t, border=_box)
        _c(ws, r, 3, a.carrier,         border=_box)
        _c(ws, r, 4, a.invoice_number,  border=_box, size=9)
        _c(ws, r, 5, a.anomaly_type,    border=_box, wrap=True)
        _c(ws, r, 6, a.description,     size=9, border=_box, wrap=True)
        _c(ws, r, 7, a.suggested_action, size=9, border=_box, wrap=True)
        _c(ws, r, 8, a.ai_explanation if a.source == "AI" else "–",
           size=9, color=_GREY, border=_box, wrap=True)
        ws.row_dimensions[r].height = 58
        r += 1
    ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════

def build_summary(data: SummaryInput, output_path: str) -> None:
    """
    Generate the 4-tab freight invoice summary and save to output_path.

    Parameters
    ----------
    data        : SummaryInput populated by your invoice parser
    output_path : full file path for the output .xlsx
    """
    carriers = sorted({i.carrier for i in data.invoices})

    carrier_totals = {c: sum(i.total_ex_vat for i in data.invoices if i.carrier == c)
                      for c in carriers}
    carrier_fuel   = {c: sum(s.amount for s in data.surcharges if s.carrier == c and s.is_fuel)
                      for c in carriers}
    carrier_other  = {c: sum(s.amount for s in data.surcharges if s.carrier == c and not s.is_fuel)
                      for c in carriers}
    carrier_frakt  = {c: sum(s.total_ex_vat for s in data.services if s.carrier == c)
                         + sum(u.amount for u in data.unallocated if u.carrier == c)
                      for c in carriers}

    wb = Workbook()
    _build_attest(wb, data, carriers, carrier_totals, carrier_fuel, carrier_other, carrier_frakt)
    _build_services(wb, data, carriers, carrier_totals, carrier_fuel, carrier_other)
    _build_surcharges(wb, data, carriers, carrier_totals)
    _build_anomalies(wb, data)
    wb.save(output_path)
    print(f"[freight_summary] Saved -> {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# EXAMPLE / SMOKE TEST  (python freight_summary.py)
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    example = SummaryInput(
        run_id="20260602_191024_dc327c83",
        generated="2026-06-02 19:11",
        files_scanned=8,
        invoices=[
            Invoice("PostNord","903110324329","2026-05-29",459.19,  "Error",
                    "Radsumma 456,18 ≠ fakturahuvud 459,19 (−3,01)",       1,"Error"),
            Invoice("PostNord","903108957122","2026-05-29",212.16,  "Warning",
                    "Inga rader kunde tolkas – ev. tilläggs-/justeringsfaktura",1,"Warning"),
            Invoice("PostNord","903109695424","2026-05-29",49954.02,"OK",
                    "Radsumma = fakturahuvud (49 954,02)",                  1,"Warning"),
            Invoice("Bring",   "4040266117", "2026-05-23",32350.65,"OK",
                    "PDF-total = specifikation (32 350,65)",                2,"Warning"),
            Invoice("Bring",   "4040267896", "2026-05-31",29660.65,"OK",
                    "PDF-total = specifikation (29 660,65)",                0,"OK"),
            Invoice("PostNord","903105199728","2026-05-22",29998.82,"OK",
                    "Radsumma = fakturahuvud (29 998,82)",                  0,"OK"),
        ],
        services=[
            # packages = real sum of antal_kolli from Bring spec (1 row per kolli, qty=1 each)
            # 389 kolli across 361 unique sändningar = 1.08 kolli/sändning average
            Service("Bring",    "Kolli (parcel)",     361, 35296.43, packages=389),
            Service("Bring",    "Pall (pallet)",        7, 12200.02, packages=8),
            # PostNord: 1 kolli = 1 sändning, packages left at 0 (kolli = shipments)
            Service("PostNord", "Kolli (parcel)",     953, 60861.67),
            Service("PostNord", "Pall (pallet)",       12,  6061.00),
            Service("PostNord", "Utlämningsställe",    29,  2211.30),
        ],
        unallocated=[
            Unallocated("PostNord", "Ej specificerat (otolkad faktura + differens)", 215.22),
        ],
        surcharges=[
            Surcharge("Bring",    "Bränsle (fuel)",  10197.74, True,  False),
            Surcharge("Bring",    "Specialhantering", 2655.00, False, False),
            Surcharge("Bring",    "Avlägset område",  1080.00, False, False),
            Surcharge("Bring",    "City",              361.00, False, False),
            Surcharge("Bring",    "Avisering",         130.00, False, False),
            Surcharge("Bring",    "Retur",              91.11, False, False),
            Surcharge("PostNord", "Bränsle (fuel)",   5369.80, True,  False),
            Surcharge("PostNord", "City",             1913.00, False, False),
            Surcharge("PostNord", "Leveransförsök",   1815.00, False, False),
            Surcharge("PostNord", "Boxadress",         600.00, False, False),
            Surcharge("PostNord", "Specialhantering",  780.00, False, False),
            Surcharge("PostNord", "Avlägset område",   512.00, False, False),
            Surcharge("PostNord", "Okänt tillägg",     158.00, False, True),   # flagged
            Surcharge("PostNord", "Valuta",             73.86, False, False),
            Surcharge("PostNord", "Svavel",             53.39, False, False),
        ],
        carrier_currency={"Bring": "NOK", "PostNord": "SEK"},
        service_surcharges=[
            # ── Bring: real surcharge amounts from actual invoice data ───────────────
            # Fuel split: 6086.32 Parcel / 4111.42 Pallet (from surcharge_lines.csv)
            ServiceSurcharge("Bring", "Kolli (parcel)", "Bränsle (fuel)",           6086.32, True,  False),
            ServiceSurcharge("Bring", "Pall (pallet)",  "Bränsle (fuel)",           4111.42, True,  False),
            # Remote Area, City, Return are Parcel-only in real data
            ServiceSurcharge("Bring", "Kolli (parcel)", "Avlägset område",          1080.00, False, False),
            ServiceSurcharge("Bring", "Kolli (parcel)", "City (storstadstillägg)",   344.00, False, False),
            ServiceSurcharge("Bring", "Kolli (parcel)", "Retur",                      91.11, False, False),
            # Special Handling, Notification, City(small) are Pallet-only in real data
            ServiceSurcharge("Bring", "Pall (pallet)",  "Specialhantering",         2655.00, False, False),
            ServiceSurcharge("Bring", "Pall (pallet)",  "City (storstadstillägg)",    17.00, False, False),
            ServiceSurcharge("Bring", "Pall (pallet)",  "Avisering",                 130.00, False, False),
            # ── PostNord: fuel (Paket prorated 953 Kolli / 29 Utlämn; Pall direct) ──
            ServiceSurcharge("PostNord", "Kolli (parcel)",   "Bränsle (fuel)",       4657.94, True,  False),
            ServiceSurcharge("PostNord", "Utlämningsställe", "Bränsle (fuel)",        141.86, True,  False),
            ServiceSurcharge("PostNord", "Pall (pallet)",    "Bränsle (fuel)",         570.00, True,  False),
            # PostNord per-shipment surcharges split proportionally by shipments (953:12:29)
            # Valuta (73.86), Svavel (53.39), Okänt (158.00) are invoice-level → not here
            ServiceSurcharge("PostNord", "Kolli (parcel)",   "City (storstadstillägg)", 1833.00, False, False),
            ServiceSurcharge("PostNord", "Pall (pallet)",    "City (storstadstillägg)",   23.00, False, False),
            ServiceSurcharge("PostNord", "Utlämningsställe", "City (storstadstillägg)",   57.00, False, False),
            ServiceSurcharge("PostNord", "Kolli (parcel)",   "Leveransförsök",          1740.00, False, False),
            ServiceSurcharge("PostNord", "Pall (pallet)",    "Leveransförsök",             22.00, False, False),
            ServiceSurcharge("PostNord", "Utlämningsställe", "Leveransförsök",             53.00, False, False),
            ServiceSurcharge("PostNord", "Kolli (parcel)",   "Boxadress",                600.00, False, False),
            ServiceSurcharge("PostNord", "Kolli (parcel)",   "Specialhantering",         748.00, False, False),
            ServiceSurcharge("PostNord", "Pall (pallet)",    "Specialhantering",           9.00, False, False),
            ServiceSurcharge("PostNord", "Utlämningsställe", "Specialhantering",          23.00, False, False),
            ServiceSurcharge("PostNord", "Kolli (parcel)",   "Avlägset område",           490.00, False, False),
            ServiceSurcharge("PostNord", "Pall (pallet)",    "Avlägset område",             6.00, False, False),
            ServiceSurcharge("PostNord", "Utlämningsställe", "Avlägset område",            16.00, False, False),
        ],
        anomalies=[
            Anomaly("Error","PostNord","903110324329","Avstämningsdiff",
                    "Radsumma 456,18 SEK ≠ fakturahuvud 459,19 SEK (−3,01).",
                    "Öppna PDF:en – leta tilläggsrad (bränsle/valuta) som inte fångats.",
                    "Differensen 3,01 SEK matchar känt PostNord-mönster – liten tilläggsrad missas vid inläsning.",
                    "Regel"),
            Anomaly("Warning","PostNord","903108957122","Inga rader tolkade",
                    "Inga radposter kunde läsas (fakturahuvud 212,16 SEK).",
                    "Verifiera manuellt om det är en kredit-/justeringsfaktura.",
                    "Sannolikt en tilläggs-, kredit- eller justeringsfaktura snarare än standard.",
                    "AI"),
            Anomaly("Warning","Bring","4040266117","Hög radkostnad",
                    "Sändning …894501 (Pall) 3 449,74 SEK överstiger pall-tröskel 2 500 SEK (+949,74).",
                    "Stäm av volym och rate mot avtal för denna tjänst.",
                    "Business Pall (9999/336) fakturerad 949,74 SEK över förväntad tröskel.",
                    "Regel"),
            Anomaly("Warning","PostNord","903109695424","Oklassat tillägg",
                    "1 tilläggsrad har okänd kategori (158 SEK).",
                    "Inspektera raden; uppdatera kategorimappningen om det återkommer.",
                    "Tillägg kunde inte matchas mot känd kategori – ev. ny PostNord-kod.",
                    "AI"),
            Anomaly("Info","Bring","4040266117","Hög kostnad/kg",
                    "Sändning …362386 1 084,20 SEK/kg – 115,5× pall-snitt (9,38 SEK/kg).",
                    "Kontrollera om fraktberäknad vikt stämmer – volymvikt kan blåsa upp debiterad vikt.",
                    "Fraktberäknad vikt 1,0 kg är debiteringsbasen. Extrem avvikare.",
                    "AI"),
        ],
    )

    build_summary(example, "summary_example.xlsx")
